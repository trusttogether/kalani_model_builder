"""Populate the blue input cells in the A.CRE template via YAML config."""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, MutableMapping, Sequence, Tuple

import openpyxl
import yaml

DEFAULT_CONFIG_PATH = Path("kalani_config.yaml")

SUMMARY_HIGHLIGHT_MAP: Sequence[tuple[str, str]] = [
    ("title", "A1"),
    ("equity_note", "A2"),
    ("irr_note", "A3"),
    ("refi_note", "A4"),
    ("distribution_note", "A5"),
]

SUMMARY_INPUT_CELL_MAP: Mapping[str, str] = {
    "property_name": "C5",
    "address": "C6",
    "city_state_zip": "C7",
    "room_count": "C9",
    "gross_square_feet": "C10",
    "analysis_start_month": "C14",
    "analysis_start_year": "C15",
    "operations_start_month": "C18",
    "operations_start_year": "C19",
    "hold_period_years": "C20",
    "mezz_toggle": "C36",
    "exit_cap_rate": "C25",
    "sale_cost_rate": "C27",
    "mezz_ltc": "C37",
    "senior_floor_rate": "C45",
    "senior_ceiling_rate": "C46",
    "perm_floor_rate": "C56",
    "perm_ceiling_rate": "C57",
    "dial_in_interest_rate": "C43",
    "secondary_ltc_control": "C48",
    "refi_interest_rate_bps": "C54",
    "operating_cashflow_pct_to_interest": "C58",
    "going_in_cap_rate": "G37",
    "ltv_ratio": "G39",
    "senior_interest_rate": "G40",
    "loan_fee_percent": "G41",
    "interest_only_months": "G42",
    "amortization_years": "G43",
}

BUDGET_GROUPS = {
    "acquisition": ("P", "Q"),
    "soft_costs": ("T", "U"),
    "hard_costs": ("X", "Y"),
    "ff_and_e": ("AB", "AC"),
    "financing_costs": ("AF", "AG"),
    "other_costs": ("AJ", "AK"),
}

WATERFALL_CELL_MAP = {
    "lp_equity_share": ("Waterfall - IRR Hurdles", "C6"),
    "tier1_lp_split": ("Waterfall - IRR Hurdles", "C9"),
    "tier1_hurdle": ("Waterfall - IRR Hurdles", "C11"),
    "tier2_gp_promote": ("Waterfall - IRR Hurdles", "C13"),
    "tier3_gp_promote": ("Waterfall - IRR Hurdles", "C18"),
    "post_stabilization_gp_share": ("Waterfall - IRR Hurdles", "C24"),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Populate blue input cells in the A.CRE template."
    )
    parser.add_argument(
        "-c",
        "--config",
        default=str(DEFAULT_CONFIG_PATH),
        help="Path to YAML config (default: kalani_config.yaml)",
    )
    parser.add_argument(
        "-t",
        "--template",
        help="Path to the template workbook (overrides config value)",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Path for the generated workbook (overrides config value)",
    )
    return parser.parse_args()


def load_config(path: Path) -> Dict:
    if not path.exists():
        raise FileNotFoundError(f"Config file not found: {path}")
    with path.open("r", encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def apply_cells(ws, mapping: Mapping[str, str], values: Mapping[str, object]) -> None:
    for friendly_name, cell in mapping.items():
        if friendly_name in values and values[friendly_name] is not None:
            ws[cell] = values[friendly_name]


def apply_summary_highlights(ws, highlights: Mapping[str, str]) -> None:
    for friendly_name, cell in SUMMARY_HIGHLIGHT_MAP:
        text = highlights.get(friendly_name)
        if text is not None:
            ws[cell] = text



def apply_budget_inputs(wb: openpyxl.Workbook, budget_cfg: Mapping[str, Mapping[str, float]]) -> None:
    if not budget_cfg:
        return
    ws = wb["Budget"]
    for group, (label_col, value_col) in BUDGET_GROUPS.items():
        entries = budget_cfg.get(group)
        if not entries:
            continue
        for label, value in entries.items():
            if value is None:
                continue
            target_row = None
            for row in range(11, 100):
                cell_value = ws[f"{label_col}{row}"].value
                if isinstance(cell_value, str) and cell_value.strip().lower() == label.strip().lower():
                    target_row = row
                    break
            if target_row is None:
                continue
            ws[f"{value_col}{target_row}"] = value


def apply_waterfall_inputs(wb: openpyxl.Workbook, waterfall_cfg: Mapping[str, float]) -> None:
    if not waterfall_cfg:
        return
    for key, (sheet_name, cell_ref) in WATERFALL_CELL_MAP.items():
        if key in waterfall_cfg and waterfall_cfg[key] is not None:
            wb[sheet_name][cell_ref] = waterfall_cfg[key]


def normalize_cell_inputs(raw: object) -> MutableMapping[str, List[Tuple[str, object]]]:
    normalized: MutableMapping[str, List[Tuple[str, object]]] = defaultdict(list)

    def add_entry(sheet: str, cell: str, value: object) -> None:
        if not sheet or not cell:
            return
        normalized[sheet].append((cell, value))

    if isinstance(raw, dict):
        for sheet, assignments in raw.items():
            if isinstance(assignments, dict):
                for cell_ref, value in assignments.items():
                    add_entry(sheet, cell_ref, value)
            elif isinstance(assignments, list):
                for entry in assignments:
                    if isinstance(entry, dict):
                        cell = entry.get("cell")
                        value = entry.get("value")
                        add_entry(sheet, cell, value)
    elif isinstance(raw, list):
        for entry in raw:
            if isinstance(entry, dict):
                sheet = entry.get("sheet")
                cell = entry.get("cell")
                value = entry.get("value")
                add_entry(sheet, cell, value)

    return normalized


def apply_cell_inputs(wb: openpyxl.Workbook, cell_inputs: MutableMapping[str, List[Tuple[str, object]]]) -> None:
    for sheet_name, entries in cell_inputs.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell_ref, value in entries:
            ws[cell_ref] = value


def apply_row_ranges(wb: openpyxl.Workbook, row_ranges: Iterable[Mapping[str, object]]) -> None:
    for entry in row_ranges or []:
        sheet_name = entry.get("sheet")
        row = entry.get("row")
        start_col = entry.get("start_col")
        end_col = entry.get("end_col")
        value = entry.get("value")
        values = entry.get("values")
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        if row is None or not start_col or not end_col:
            continue
        ws = wb[sheet_name]
        try:
            start_idx = openpyxl.utils.column_index_from_string(start_col)
            end_idx = openpyxl.utils.column_index_from_string(end_col)
        except ValueError:
            continue
        if start_idx > end_idx:
            start_idx, end_idx = end_idx, start_idx
        seq = values if values else [value for _ in range(end_idx - start_idx + 1)]
        if not seq:
            continue
        for offset, col_idx in enumerate(range(start_idx, end_idx + 1)):
            if offset >= len(seq):
                break
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws[f"{col_letter}{row}"] = seq[offset]


def main() -> None:
    args = parse_args()
    config = load_config(Path(args.config))

    project_cfg = config.get("project", {})
    template_default = project_cfg.get("template", "A.CRE-Hotel-Development-Model-beta-v1.57.xlsx")
    output_default = project_cfg.get("output", "updated_model.xlsx")
    template_path = Path(args.template) if args.template else Path(template_default)
    output_path = Path(args.output) if args.output else Path(output_default)
    if not template_path.exists():
        raise FileNotFoundError(f"Cannot find template workbook at {template_path}")

    wb = openpyxl.load_workbook(template_path)

    summary_cfg = config.get("summary") or {}
    summary_sheet = wb["Summary"]
    apply_summary_highlights(summary_sheet, summary_cfg.get("highlights", {}))
    apply_cells(summary_sheet, SUMMARY_INPUT_CELL_MAP, summary_cfg.get("inputs", {}))

    apply_budget_inputs(wb, config.get("budget") or {})
    apply_waterfall_inputs(wb, config.get("waterfall") or {})

    manual_cells = normalize_cell_inputs(config.get("manual_cells"))
    cell_inputs = normalize_cell_inputs(config.get("cell_inputs"))
    for sheet, entries in manual_cells.items():
        cell_inputs[sheet].extend(entries)
    apply_cell_inputs(wb, cell_inputs)
    apply_row_ranges(wb, config.get("row_ranges"))

    wb.save(output_path)
    print(f"Updated workbook saved to {output_path.resolve()}")


if __name__ == "__main__":
    main()
