import openpyxl

# Load the original workbook

wb = openpyxl.load_workbook('A.CRE-Hotel-Development-Model-beta-v1.57.xlsx')

# 1. Update Summary Sheet

summary_sheet = wb['Summary']

summary_sheet['A1'] = "Kalani Resort Updated Model - Nov 22, 2025"

summary_sheet['A2'] = "Total Equity Needed: $2,090,000"

summary_sheet['A3'] = "Projected IRR: 25-35%"

summary_sheet['A4'] = "Return of Capital: End 2026 via Refinance (Month 24)"

summary_sheet['A5'] = "Expected Returns: 150% Preferred to Class B/C, then 30% ongoing"

# 2. Update AnnualCF Sheet with Projections

annual_cf_sheet = wb['AnnualCF']

# Assuming row 6 is headers like Year 0,1,2,... Update Years 1-3 (2025-2027)

# Find or assume rows for revenue, expenses, NOI (adjust if labels differ; e.g., row 10 for revenue)

# For simplicity, updating placeholder rows - inspect your sheet and adjust indices if needed

# 2025 (Year 1)

annual_cf_sheet.cell(row=10, column=4).value = 5178540  # Room Revenue (assuming col D=Year1)

annual_cf_sheet.cell(row=11, column=4).value = 1202208  # Ancillary

annual_cf_sheet.cell(row=12, column=4).value = 6382248  # Total Gross

annual_cf_sheet.cell(row=13, column=4).value = 5290082  # Expenses

annual_cf_sheet.cell(row=14, column=4).value = 1092166  # NOI

# 2026 (Year 2)

annual_cf_sheet.cell(row=10, column=5).value = 6417132

annual_cf_sheet.cell(row=11, column=5).value = 1412778

annual_cf_sheet.cell(row=12, column=5).value = 7829910

annual_cf_sheet.cell(row=13, column=5).value = 6406392

annual_cf_sheet.cell(row=14, column=5).value = 1423518

# 2027 (Year 3)

annual_cf_sheet.cell(row=10, column=6).value = 7058845

annual_cf_sheet.cell(row=11, column=6).value = 1554056

annual_cf_sheet.cell(row=12, column=6).value = 8612901

annual_cf_sheet.cell(row=13, column=6).value = 6918903

annual_cf_sheet.cell(row=14, column=6).value = 1693998

# 3. Add New Sheet: Investor Returns

investor_sheet = wb.create_sheet('Investor Returns')

# High-Level Metrics Table (A1:B7)

investor_sheet['A1'] = "Metric"

investor_sheet['B1'] = "Details"

investor_sheet['A2'] = "Investor Return Schedule"

investor_sheet['A3'] = "Total Equity Raised"

investor_sheet['B3'] = "$2,090,000"

investor_sheet['A4'] = "Projected IRR"

investor_sheet['B4'] = "25-35%"

investor_sheet['A5'] = "Return of Capital Timeline"

investor_sheet['B5'] = "End of 2026 (Month 24 post-closing) via Refinance - Full capital recovery + 50% profit (150% total return on invested capital)"

investor_sheet['A6'] = "Preferred Return to Class B/C"

investor_sheet['B6'] = "100% Capital Recovery + 50% Profit (150% Total)"

investor_sheet['A7'] = "Ongoing Distributions Post-Refi"

investor_sheet['B7'] = "70% to Class A, 30% to Class B/C based on NOI distributions"

investor_sheet['A8'] = "Exit Scenario (2027 Sale/Refi)"

investor_sheet['B8'] = "Stabilized Value $21.1M, Net Proceeds after Debt ~$18M, Potential 8-10x Multiple on Equity"

# Annual Distributions Schedule (A10:C14)

investor_sheet['A10'] = "Year"

investor_sheet['B10'] = "Estimated Distributions ($)"

investor_sheet['C10'] = "Cumulative Return ($)"

investor_sheet['A11'] = "2025"

investor_sheet['B11'] = "$0"

investor_sheet['C11'] = "$0"

investor_sheet['A12'] = "2026"

investor_sheet['B12'] = "$3,135,000"

investor_sheet['C12'] = "$3,135,000"

investor_sheet['A13'] = "2027"

investor_sheet['B13'] = "$508,199"

investor_sheet['C13'] = "$3,643,199"

investor_sheet['A14'] = "Post-2027 (Annual)"

investor_sheet['B14'] = "30% of Annual NOI (~$508k base, growing)"

investor_sheet['C14'] = "Ongoing"

# Save the updated workbook

wb.save('updated_kalani_model_v2.xlsx')

print("Updated workbook saved as 'updated_kalani_model_v2.xlsx'!")